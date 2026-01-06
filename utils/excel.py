"""
Excel utilities for reading and writing ranking data
Preserves Excel format by using openpyxl directly
"""

import pandas as pd
from typing import Optional
from openpyxl import load_workbook


class ExcelHandler:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.keyword_column = "Local Keyword Ideas"
        self.places_column = "Google Places"
        self.links_column = "Google Links"
    
    def read_keywords(self, sheet_name: str = "Keywords", header_row: int = 1) -> pd.DataFrame:
        """
        Read keywords from Excel file using pandas (for easy filtering)
        
        Args:
            sheet_name: Name of the sheet to read
            header_row: Row number (0-indexed) containing headers
            
        Returns:
            DataFrame with keywords
        """
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=header_row)
            print(f"✓ Loaded {len(df)} rows from {self.excel_file}")
            return df
        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}")
    
    def update_rankings(
        self,
        row_index: int,
        places_rank: Optional[int],
        links_rank: Optional[int],
        sheet_name: str = "Keywords",
        header_row: int = 1
    ) -> bool:
        """
        Update rankings directly in Excel file preserving format
        
        Args:
            row_index: Pandas DataFrame index (not Excel row number)
            places_rank: Google Places rank (or None)
            links_rank: Google Links rank (or None)
            sheet_name: Name of the sheet
            header_row: Row number (0-indexed) containing headers
            
        Returns:
            True if successful
        """
        try:
            # Load workbook
            wb = load_workbook(self.excel_file)
            if sheet_name not in wb.sheetnames:
                print(f"  ✗ Sheet '{sheet_name}' not found")
                return False
            
            ws = wb[sheet_name]
            
            # Excel row number for headers (header_row is 0-indexed, Excel rows start at 1)
            header_row_num = header_row + 1
            
            # Find column letters for Google Places and Google Links
            places_col_letter = None
            links_col_letter = None
            
            # Search in header row and next row (for merged cells like "Google Rank")
            for search_row in [header_row_num, header_row_num + 1]:
                for cell in ws[search_row]:
                    if cell.value:
                        cell_value = str(cell.value).strip().lower()
                        if self.places_column.lower() in cell_value and not places_col_letter:
                            places_col_letter = cell.column_letter
                        if self.links_column.lower() in cell_value and not links_col_letter:
                            links_col_letter = cell.column_letter
            
            if not places_col_letter:
                print(f"  ✗ Column '{self.places_column}' not found")
                return False
            if not links_col_letter:
                print(f"  ✗ Column '{self.links_column}' not found")
                return False
            
            # Calculate Excel row number (pandas row_index + header_row + 2)
            # If header_row=1 (row 2 in Excel), data starts at row 3
            excel_row = row_index + header_row + 2
            
            # Update cells with numeric values or "Not Found"
            if places_rank is not None:
                ws[f"{places_col_letter}{excel_row}"] = places_rank
                print(f"  ✓ Updated {self.places_column} (row {excel_row}): {places_rank}")
            else:
                ws[f"{places_col_letter}{excel_row}"] = "Not Found"
                print(f"  ✓ Updated {self.places_column} (row {excel_row}): Not Found")
            
            if links_rank is not None:
                ws[f"{links_col_letter}{excel_row}"] = links_rank
                print(f"  ✓ Updated {self.links_column} (row {excel_row}): {links_rank}")
            else:
                ws[f"{links_col_letter}{excel_row}"] = "Not Found"
                print(f"  ✓ Updated {self.links_column} (row {excel_row}): Not Found")
            
            # Save workbook (updates existing file)
            wb.save(self.excel_file)
            return True
            
        except Exception as e:
            print(f"  ✗ Error updating Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
